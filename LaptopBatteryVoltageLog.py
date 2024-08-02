import psutil
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

def get_battery_voltage():
    battery = psutil.sensors_battery()
    voltage = battery.percent * 0.01 * 12.6  # Assuming full charge is 12.6 volts
    return voltage

def log_battery_voltage(file_name):
    try:
        workbook = load_workbook(file_name)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Date", "Time", "Battery Voltage (V)"])
    
    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H:%M:%S")
    voltage = get_battery_voltage()
    
    sheet.append([date_str, time_str, voltage])
    
    workbook.save(file_name)

if __name__ == "__main__":
    file_name = "battery_log.xlsx"
    log_battery_voltage(file_name)
    print("Battery voltage logged successfully.")
